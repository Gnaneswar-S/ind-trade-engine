"""
report_service.py
──────────────────────────────────────────────────────────────
Export trade analysis results to formatted Excel (.xlsx) and PDF.

Libraries:
  Excel  → openpyxl
  PDF    → reportlab

Public API:
  export_to_excel(result, product, mode, user_email) → BytesIO
  export_to_pdf  (result, product, mode, user_email) → BytesIO
  get_report_filename(product, mode, ext)            → str
──────────────────────────────────────────────────────────────
"""

import io
import re
from datetime import datetime, timezone
from typing import Any

# ── Excel ──────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ── PDF ────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT


# ──────────────────────────────────────────────
# CONSTANTS — colour palette
# ──────────────────────────────────────────────

class _Pal:
    NAVY       = "1A3A5C"
    BLUE       = "1A6FA8"
    LIGHT_BLUE = "D6EAF8"
    WHITE      = "FFFFFF"
    DARK_GRAY  = "2D3748"
    MID_GRAY   = "718096"
    LIGHT_GRAY = "F7FAFC"
    ALT_ROW    = "EBF5FB"
    GREEN      = "1E8449"
    AMBER      = "D68910"
    RED        = "C0392B"
    WARN_BG    = "FEF9E7"
    WARN_BORDER= "D4AC0D"


# Human-readable label + display category for every possible result key
_FIELD_META: dict[str, tuple[str, str]] = {
    "hs_code":                          ("HS Code (ITC-HS 8-digit)",         "id"),
    "product_description":              ("Product Description",               "text"),
    "basic_customs_duty_percent":       ("Basic Customs Duty (BCD)",          "percent"),
    "social_welfare_surcharge_percent": ("Social Welfare Surcharge (SWS)",    "percent"),
    "igst_percent":                     ("IGST",                              "percent"),
    "total_landed_cost_percent":        ("Total Landed Cost (approx.)",       "percent"),
    "import_policy_status":             ("Import Policy Status",              "status"),
    "license_required":                 ("Licence Required",                  "bool"),
    "scomet_applicable":                ("SCOMET Applicable",                 "bool"),
    "special_conditions":               ("Special Conditions",                "text"),
    "export_policy_status":             ("Export Policy Status",              "status"),
    "rodtep_applicable":                ("RoDTEP Applicable",                 "bool"),
    "rodtep_rate_percent":              ("RoDTEP Rate",                       "percent"),
    "rosctl_applicable":                ("RoSCTL Applicable",                 "bool"),
    "export_duty_percent":              ("Export Duty",                       "percent"),
    "export_incentive_notes":           ("Export Incentive Notes",            "text"),
    "restricted_countries":             ("Country Restrictions",              "text"),
    "documentation_required":           ("Documentation Required",            "text"),
    "gst_percent":                      ("GST Rate",                          "percent"),
    "gst_category":                     ("GST Category",                      "text"),
    "itc_available":                    ("ITC Available",                     "bool"),
    "itc_conditions":                   ("ITC Conditions",                    "text"),
    "compliance_requirements":          ("Compliance Requirements",           "text"),
    "fssai_required":                   ("FSSAI Required",                    "bool"),
    "bis_required":                     ("BIS Required",                      "bool"),
    "other_regulatory":                 ("Other Regulatory Requirements",     "text"),
    "risk_flags":                       ("Risk Flags",                        "risk"),
    "data_confidence":                  ("AI Data Confidence",                "status"),
    "validation_warning":               ("⚠ Validation Warning",             "warn"),
    "note":                             ("Classification Note",               "text"),
    "chapter":                          ("HS Chapter",                        "text"),
    "confidence":                       ("Classification Confidence",         "text"),
}

# Internal prefixed keys or metadata keys to skip
_SKIP_KEYS = {"error", "raw_response"}


# ──────────────────────────────────────────────
# SHARED HELPERS
# ──────────────────────────────────────────────

def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")


def _ts_compact() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")


def _status_color_xl(value: str) -> str:
    """Return hex color string for Excel cell font."""
    v = str(value).lower()
    if any(x in v for x in ["free", "true", "yes", "high", "applicable"]):
        return _Pal.GREEN
    if any(x in v for x in ["restricted", "medium", "conditional"]):
        return _Pal.AMBER
    if any(x in v for x in ["prohibited", "false", "no", "low", "null"]):
        return _Pal.RED
    return _Pal.DARK_GRAY


def _bool_label(value: Any) -> str:
    return "✅  Yes" if str(value).lower() in ("true", "1", "yes") else "❌  No"


def _build_rows(result: dict) -> list[tuple[str, str, str]]:
    """
    Convert result dict → list of (label, value_str, field_type).
    Skips internal keys and blank values.
    """
    rows = []
    for key, value in result.items():
        if key.startswith("_") or key in _SKIP_KEYS:
            continue
        if value is None or str(value).strip() in ("", "null", "none", "None"):
            continue
        label, ftype = _FIELD_META.get(key, (key.replace("_", " ").title(), "text"))
        str_val = _bool_label(value) if ftype == "bool" else str(value)
        rows.append((label, str_val, ftype))
    return rows


# ──────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────

def _xl_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _xl_font(bold=False, color=_Pal.DARK_GRAY, size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _xl_border() -> Border:
    thin = Side(style="thin", color="D8E4F0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _xl_align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def export_to_excel(
    result: dict,
    product: str,
    mode: str,
    user_email: str,
) -> io.BytesIO:
    """
    Build a polished, formatted Excel (.xlsx) report.
    Returns BytesIO — pass directly to st.download_button.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{mode} Report"

    # Column widths
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 16

    # ── ROW 1: Banner ──────────────────────────────
    ws.merge_cells("A1:C1")
    banner = ws["A1"]
    banner.value     = "🇮🇳  Indian Trade Intelligence Engine"
    banner.font      = _xl_font(bold=True, color=_Pal.WHITE, size=15)
    banner.fill      = _xl_fill(_Pal.NAVY)
    banner.alignment = _xl_align("center", "center")
    ws.row_dimensions[1].height = 44

    # ── ROWS 2-6: Metadata ─────────────────────────
    meta = [
        ("Report Type",  f"{mode} Analysis Report"),
        ("Product",       product[:120]),
        ("Generated By",  user_email),
        ("Generated At",  _now_str()),
        ("Data Source",   "Llama 3.3 70B via NVIDIA API · ITC-HS India"),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        ws.row_dimensions[i].height = 18
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        lc.font = _xl_font(bold=True, color=_Pal.NAVY, size=9)
        lc.fill = _xl_fill(_Pal.LIGHT_BLUE)
        lc.alignment = _xl_align("left")
        lc.border    = _xl_border()
        vc.font = _xl_font(size=9)
        vc.alignment = _xl_align("left", wrap=True)
        vc.border    = _xl_border()
        # Merge value across B+C
        ws.merge_cells(f"B{i}:C{i}")

    # ── ROW 8: Section header ──────────────────────
    SECTION_ROW = len(meta) + 3
    ws.merge_cells(f"A{SECTION_ROW}:C{SECTION_ROW}")
    sh = ws.cell(row=SECTION_ROW, column=1,
                 value=f"  📊  {mode.upper()} ANALYSIS RESULTS")
    sh.font      = _xl_font(bold=True, color=_Pal.WHITE, size=11)
    sh.fill      = _xl_fill(_Pal.BLUE)
    sh.alignment = _xl_align("left", "center")
    ws.row_dimensions[SECTION_ROW].height = 28

    # ── ROW 9: Column headers ──────────────────────
    HDR_ROW = SECTION_ROW + 1
    ws.row_dimensions[HDR_ROW].height = 20
    for col, text in enumerate(["Field", "Value", "Status"], start=1):
        c = ws.cell(row=HDR_ROW, column=col, value=text)
        c.font      = _xl_font(bold=True, color=_Pal.WHITE, size=9)
        c.fill      = _xl_fill(_Pal.DARK_GRAY)
        c.alignment = _xl_align("center")
        c.border    = _xl_border()

    # ── DATA ROWS ──────────────────────────────────
    data_rows = _build_rows(result)
    for idx, (label, str_val, ftype) in enumerate(data_rows):
        r = HDR_ROW + 1 + idx
        ws.row_dimensions[r].height = 20
        bg = _xl_fill(_Pal.ALT_ROW if idx % 2 == 0 else _Pal.WHITE)

        # A: field label
        a = ws.cell(row=r, column=1, value=label)
        a.font = _xl_font(bold=True, color=_Pal.NAVY, size=9)
        a.fill = bg
        a.alignment = _xl_align("left", "center")
        a.border    = _xl_border()

        # B: value
        b = ws.cell(row=r, column=2, value=str_val)
        b.font = _xl_font(size=9)
        b.fill = bg
        b.alignment = _xl_align("left", "center", wrap=True)
        b.border    = _xl_border()

        # C: status indicator
        if ftype == "bool":
            indicator = str_val
            c_color   = _Pal.GREEN if "Yes" in str_val else _Pal.RED
        elif ftype == "status":
            indicator = str_val.upper()
            c_color   = _status_color_xl(str_val)
        elif ftype in ("percent", "id"):
            indicator = "—"
            c_color   = _Pal.MID_GRAY
        elif ftype in ("risk", "warn"):
            indicator = "⚠ Review"
            c_color   = _Pal.AMBER
        else:
            indicator = "—"
            c_color   = _Pal.MID_GRAY

        c = ws.cell(row=r, column=3, value=indicator)
        c.font      = _xl_font(bold=True, color=c_color, size=9)
        c.fill      = bg
        c.alignment = _xl_align("center")
        c.border    = _xl_border()

    # ── FOOTER ─────────────────────────────────────
    foot_row = HDR_ROW + len(data_rows) + 2
    ws.merge_cells(f"A{foot_row}:C{foot_row}")
    f = ws.cell(
        row=foot_row, column=1,
        value="⚠  AI-generated report. Verify with official sources: "
              "icegate.gov.in · DGFT FTP · CBIC Customs Tariff. Not legal advice.",
    )
    f.font      = _xl_font(color=_Pal.MID_GRAY, size=8)
    f.fill      = _xl_fill(_Pal.LIGHT_BLUE)
    f.alignment = _xl_align("center", "center", wrap=True)
    ws.row_dimensions[foot_row].height = 28

    # Freeze panes below metadata
    ws.freeze_panes = f"A{SECTION_ROW}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────
# PDF EXPORT
# ──────────────────────────────────────────────

def export_to_pdf(
    result: dict,
    product: str,
    mode: str,
    user_email: str,
) -> io.BytesIO:
    """
    Build a polished A4 PDF report.
    Returns BytesIO — pass directly to st.download_button.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.8 * cm,
        rightMargin=1.8 * cm,
        topMargin=2.0 * cm,
        bottomMargin=2.0 * cm,
    )

    base = getSampleStyleSheet()

    # Custom paragraph styles
    S = {
        "title": ParagraphStyle(
            "s_title",
            parent=base["Title"],
            fontSize=18, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1A3A5C"),
            spaceAfter=4, alignment=TA_CENTER,
        ),
        "subtitle": ParagraphStyle(
            "s_sub",
            parent=base["Normal"],
            fontSize=9, fontName="Helvetica",
            textColor=colors.HexColor("#718096"),
            spaceAfter=2, alignment=TA_CENTER,
        ),
        "section": ParagraphStyle(
            "s_sec",
            parent=base["Normal"],
            fontSize=10, fontName="Helvetica-Bold",
            textColor=colors.white,
            backColor=colors.HexColor("#1A6FA8"),
            spaceBefore=10, spaceAfter=0,
            leftIndent=0, borderPad=7,
        ),
        "meta_key": ParagraphStyle(
            "s_mk",
            parent=base["Normal"],
            fontSize=8.5, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1A3A5C"),
        ),
        "meta_val": ParagraphStyle(
            "s_mv",
            parent=base["Normal"],
            fontSize=8.5, fontName="Helvetica",
            textColor=colors.HexColor("#2D3748"),
        ),
        "field_key": ParagraphStyle(
            "s_fk",
            parent=base["Normal"],
            fontSize=8.5, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1A3A5C"),
        ),
        "footer": ParagraphStyle(
            "s_foot",
            parent=base["Normal"],
            fontSize=7, fontName="Helvetica-Oblique",
            textColor=colors.HexColor("#718096"),
            alignment=TA_CENTER,
        ),
    }

    def _val_style(ftype: str, value: str) -> ParagraphStyle:
        """Value cell style — colour-coded by semantic type."""
        color_map = {
            "status": _status_color_pdf(value),
            "bool":   ("#1E8449" if "Yes" in value else "#C0392B"),
            "risk":   "#C0392B",
            "warn":   "#D68910",
        }
        hex_c = color_map.get(ftype, "#2D3748")
        return ParagraphStyle(
            f"dyn_{ftype}",
            parent=base["Normal"],
            fontSize=8.5, fontName="Helvetica-Bold" if ftype in ("status", "bool") else "Helvetica",
            textColor=colors.HexColor(hex_c),
        )

    story = []

    # ── TITLE ──────────────────────────────────────
    story.append(Paragraph("🇮🇳  Indian Trade Intelligence Engine", S["title"]))
    story.append(Paragraph(f"{mode} Analysis Report", S["subtitle"]))
    story.append(HRFlowable(
        width="100%", thickness=2,
        color=colors.HexColor("#1A6FA8"), spaceAfter=12,
    ))

    # ── METADATA TABLE ─────────────────────────────
    meta = [
        ("Report Type",  f"{mode} Analysis"),
        ("Product",       product[:100]),
        ("Generated By",  user_email),
        ("Generated At",  _now_str()),
        ("Data Source",   "Llama 3.3 70B via NVIDIA API"),
    ]
    meta_table_data = [
        [Paragraph(k, S["meta_key"]), Paragraph(v, S["meta_val"])]
        for k, v in meta
    ]
    meta_tbl = Table(meta_table_data, colWidths=[4.5 * cm, 13 * cm])
    meta_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (0, -1), colors.HexColor("#D6EAF8")),
        ("BACKGROUND",    (1, 0), (1, -1), colors.HexColor("#F7FAFC")),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#BFD7ED")),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 14))

    # ── RESULTS SECTION HEADER ─────────────────────
    story.append(Paragraph(f"  {mode.upper()} ANALYSIS RESULTS", S["section"]))
    story.append(Spacer(1, 6))

    # ── RESULTS TABLE ──────────────────────────────
    data_rows = _build_rows(result)
    if data_rows:
        tbl_data = []
        for idx, (label, str_val, ftype) in enumerate(data_rows):
            tbl_data.append([
                Paragraph(label, S["field_key"]),
                Paragraph(str_val[:220], _val_style(ftype, str_val)),
            ])

        result_tbl = Table(tbl_data, colWidths=[6 * cm, 11.5 * cm])
        row_styles = []
        for idx in range(len(tbl_data)):
            bg = colors.HexColor("#EBF5FB") if idx % 2 == 0 else colors.white
            row_styles.append(("BACKGROUND", (0, idx), (-1, idx), bg))

        result_tbl.setStyle(TableStyle([
            ("GRID",          (0, 0), (-1, -1), 0.35, colors.HexColor("#BFD7ED")),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            *row_styles,
        ]))
        story.append(KeepTogether(result_tbl))

    # ── VALIDATION WARNING ─────────────────────────
    if result.get("validation_warning"):
        story.append(Spacer(1, 10))
        warn_tbl = Table(
            [[Paragraph(
                f"⚠  {result['validation_warning']}",
                ParagraphStyle("s_warn", parent=base["Normal"],
                               fontSize=8.5, fontName="Helvetica-Bold",
                               textColor=colors.HexColor("#D68910"))
            )]],
            colWidths=[17.5 * cm],
        )
        warn_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#FEF9E7")),
            ("BOX",           (0, 0), (-1, -1), 1, colors.HexColor("#D4AC0D")),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ]))
        story.append(warn_tbl)

    # ── DISCLAIMER ─────────────────────────────────
    story.append(Spacer(1, 22))
    story.append(HRFlowable(
        width="100%", thickness=0.5,
        color=colors.HexColor("#BFD7ED"), spaceAfter=6,
    ))
    story.append(Paragraph(
        "This report is AI-generated for informational purposes only. "
        "Always verify with official sources: icegate.gov.in · DGFT FTP · CBIC Customs Tariff. "
        "Not legal or financial advice.",
        S["footer"],
    ))
    story.append(Paragraph(
        f"Generated by Trade Intelligence Engine · {_now_str()}",
        S["footer"],
    ))

    doc.build(story)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────

def _status_color_pdf(value: str) -> str:
    v = str(value).lower()
    if any(x in v for x in ["free", "true", "yes", "high", "applicable"]):
        return "#1E8449"
    if any(x in v for x in ["restricted", "medium"]):
        return "#D68910"
    if any(x in v for x in ["prohibited", "false", "no", "low"]):
        return "#C0392B"
    return "#2D3748"


def get_report_filename(product: str, mode: str, ext: str) -> str:
    """Generate a clean, timestamped filename."""
    safe = re.sub(r"[^a-zA-Z0-9 \-_]", "_", product[:40]).strip().replace(" ", "_")
    return f"TradeReport_{mode}_{safe}_{_ts_compact()}.{ext}"